#!/usr/bin/env python3
"""
生成检查规则Excel文件（两张Sheet：PDF版规则 / Word版规则）

运行方式：
    python generate_rules_excel.py

输出文件：检查规则汇总.xlsx

注意：PDF版检查对象为NSFC国家自然科学基金申请书（91条规则：初筛+LLM质量+参考文献验证）；
Word版检查对象为大学生毕业论文（仅60条参考文献GB/T 7714验证规则）。

规则更改时，请同步更新本文件中的 PRELIMINARY_RULES、LLM_QUALITY_RULES 和 REFERENCE_RULES 列表。
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==================== 规则数据定义 ====================
# 每条规则: (序号, 规则名, 检查位置, 实现方式, 正则/LLM提示词摘要, 严重度, 简要说明, 数据来源)

PRELIMINARY_RULES = [
    # ---- 原始正则规则 1~8 ----
    ("1", "申请代码格式检查", "基本信息表", "正则",
     r"r'申请代码.*?[:：]\s*(\d{4})'  # 匹配4位数字申请代码",
     "high", "第一个申请代码必须是4位纯数字",
     "完整MD（含HTML表格）"),

    ("2", "合作单位规范检查", "基本信息表", "正则",
     r"r'合作.*?单位.*?[:：]?\s*(.+)'  # 提取合作单位名称，检查是否含附院关键词",
     "medium", "合作单位不应包含附属医院/附院（附院应列为依托单位）",
     "完整MD（纯文本）"),

    ("3", "主要参与者人数统计", "基本信息表/参与者表格", "正则",
     r"r'(\d+)\s*人'  # 提取声明人数，与表格实际行数对比",
     "high", "声明的参与者人数与实际表格行数是否一致",
     "完整MD（纯文本+表格行）"),

    ("4.1", "报告正文最新模板章节检查", "报告正文", "正则",
     r"r'立项依据|研究内容|研究基础'  # 检查新模板必需章节",
     "high", "新模板必须包含：立项依据、研究内容、研究基础",
     "完整MD（纯文本）"),

    ("4.2", "报告正文空白页检查", "报告正文", "正则",
     r"r'\n{5,}'  # 连续≥5个换行视为空白页",
     "medium", "检测正文中是否存在空白页（连续空行）",
     "完整MD（纯文本）"),

    ("4.3", "H18专项特殊内容检查", "报告正文（立项依据前）", "正则",
     r"r'H18'检查申请代码 + r'临床实践|临床问题|临床需求'",
     "high", "H18面上项目须在立项依据前说明临床实践来源",
     "完整MD（含HTML表格）"),

    ("4.4", "容易漏写章节检查", "报告正文", "正则",
     r"r'可行性分析|预期研究结果|技术路线'  # 检查常见漏写章节",
     "medium", "提醒检查：可行性分析、预期研究结果、技术路线图",
     "完整MD（纯文本）"),

    ("5.1", "年度计划与执行期时间一致性", "年度研究计划", "正则",
     r"r'(\d{4})\s*[年]\s*(\d{1,2})\s*[月]?.*?[-—~至].*?(\d{4})\s*[年]\s*(\d{1,2})\s*[月]?'  # 支持【年度研究计划】等多种标题格式",
     "high", "年度计划首尾时间必须与项目执行期一致（支持【】括号包裹的章节标题）",
     "完整MD（含HTML表格）"),

    ("5.2", "年度计划时间衔接检查", "年度研究计划", "正则",
     r"前一期结束日期 + timedelta(days=1) == 下一期开始日期",
     "high", "相邻年度间不允许缺少或重叠任何天数",
     "完整MD（含HTML表格）"),

    ("6", "联系方式格式检查（电话）", "基本信息表", "正则",
     r"r'1[3-9]\d{9}'  # 11位手机号 或 固话格式",
     "low", "手机号11位或固话格式检查",
     "完整MD（纯文本）"),

    ("7", "联系方式格式检查（邮箱）", "基本信息表", "正则",
     r"r'[\w\.\-]+@[\w\.\-]+\.\w+'  # 邮箱格式正则",
     "low", "邮箱需符合 xxx@xxx.xxx 格式",
     "完整MD（纯文本）"),

    ("8", "组织机构代码检查", "基本信息表", "正则",
     r"r'[0-9A-Z]{18}'  # 统一社会信用代码18位",
     "low", "依托单位统一社会信用代码应为18位数字/大写字母",
     "完整MD（纯文本）"),

    # ---- 新增正则规则 A1~A10 ----
    ("A1", "预算总额与分项合计一致性", "经费申请表", "正则",
     r"提取含金额关键词行 → 数值求和 → 与'合计'行对比，允许0.01万元误差",
     "high", "分项预算求和必须等于合计金额",
     "完整MD（含HTML表格）"),

    ("A2", "项目执行期限合理性", "基本信息表", "正则",
     r"r'研究期限|执行期限'提取起止年月 → 计算跨度≤5年 & 起始年≥当前年",
     "high/medium", "执行期≤5年；起始年份不应早于当前年",
     "完整MD（含HTML表格）"),

    ("A3", "身份证号格式校验", "基本信息表/参与者", "正则",
     r"r'\d{17}[\dXx]' 提取18位 → 校验位算法(权重因子 7,9,10,5,8,4,2,1,6,3,7,9,10,5,8,4,2)",
     "high", "18位身份证号校验位验算",
     "完整MD（纯文本）"),

    ("A4", "参与者工作月数合理性", "主要参与者信息表", "正则",
     r"定位含'工作.*月|投入.*月|每年.*月'列头（排除'出生年月'列）→ 提取每行数值 → 判断≤12",
     "high", "每人每年投入工作月数不得超过12个月（自动排除“出生年月”列避免误读）",
     "完整MD（纯文本+表格行）"),

    ("A5", "关键词数量检查", "基本信息", "正则",
     r"r'关键词.*?[:：]\s*(.+)' → 按;；,，、分割 → 检查3≤数量≤5",
     "medium", "NSFC要求3~5个关键词",
     "完整MD（含HTML表格）"),

    ("A6", "参考文献序号连续性", "参考文献", "正则",
     r"r'^\s*\[(\d+)\]' 提取所有编号 → 检查从[1]连续递增",
     "high", "参考文献编号必须从[1]开始且连续",
     "参考文献区域MD"),

    ("A7", "参考文献引用覆盖", "参考文献/正文", "正则",
     r"正文中\[(\d+)\]集合 vs 文末\[(\d+)\]集合 → 双向差集",
     "high/medium", "文末文献须在正文引用，正文引用须在文末有条目",
     "正文MD + 参考文献区域MD"),

    ("A8", "摘要字数限制", "项目摘要/英文摘要", "正则",
     r"中文: len(re.findall(r'[\u4e00-\u9fa5]'))≤400；英文: len(words)≤500",
     "medium", "中文摘要≤400字，英文摘要≤500词",
     "完整MD（含HTML表格）"),

    ("A9", "项目名称中英文一致性", "基本信息", "正则",
     r"中文字数 vs 英文词数 → 比值范围 0.3~4.0",
     "medium", "中英文项目名称长度比率应在合理范围内",
     "完整MD（含HTML表格）"),

    ("A10", "正文重复段落检测", "报告正文", "正则",
     r"段落去空白后≥50字符 → 滑动窗口(50,step=10) → 子串匹配检测重复（排除参考文献章节）",
     "medium", "检测正文中≥50字符的重复段落（自动排除参考文献区域）",
     "完整MD（纯文本）"),

    ("A11", "基本信息必填字段检查", "基本信息表", "正则",
     r"定位表格中'主要研究领域'等必填字段的<td> → 检查紧邻的下一个<td>是否为空",
     "high", "检查关键必填字段（如“主要研究领域”）是否留空未填",
     "完整MD（含HTML表格）"),

    ("A12", "资助类别金额限制", "基本信息表", "正则",
     r"提取'资助类别'与'申请经费' → 青年≤30/面上≤80/地区≤40",
     "high", "检查青年/面上/地区项目的申请经费是否超标",
     "完整MD（含HTML表格）"),

    ("A13", "资助类别年限限制", "基本信息表", "正则",
     r"提取'资助类别'与'执行期限' → 青年≤3年/面上≤4年/地区≤4年",
     "high", "检查青年/面上/地区项目的执行年限是否超标",
     "完整MD（含HTML表格）"),

    ("A14", "总人数统计校验", "基本信息表/人员统计", "正则",
     r"提取'总人数'与各分项(高级/中级...) → 校验 总人数 == Sum(分项)",
     "high", "人员统计表中总人数必须等于各职称人数之和",
     "完整MD（含HTML表格）"),

    # ---- 混合规则 C1~C2 ----
    ("C1", "参考文献时效性检查", "参考文献", "混合(正则+LLM)",
     r"正则提取每条文献出版年 → 近5年占比<30%告警 → ≥20年老旧文献提示",
     "medium/low", "近5年文献占比应≥30%；提示超过20年的老旧文献",
     "参考文献区域MD"),

    ("C2", "参考文献自引率检查", "参考文献/基本信息表", "混合(正则+LLM)",
     r"正则提取申请人姓名+文献作者 → 中文名直接匹配 → 英文名不确定时调用一次LLM判断\n"
     r"LLM prompt: '判断以下参考文献的作者中是否包含\"{申请人}\"。仅回复JSON数组。'",
     "medium/low", "自引率>40%告警，>30%低优先级提示；英文姓名匹配时可选LLM辅助",
     "参考文献区域MD + 基本信息MD/HTML表格"),
]

LLM_QUALITY_RULES = [
    ("LLM-1", "4.1.1 立项依据质量评估", "报告正文（立项依据）", "LLM",
     "LLM system prompt中: '评估是否说清楚为什么要开展此项研究，研究的价值何在'\n"
     "状态: passed / failed / not_applicable",
     "high", "评估立项依据是否说清研究动机与价值",
     "MD分章节文本 → LLM"),

    ("LLM-2", "4.1.2 研究内容质量评估", "报告正文（研究内容）", "LLM",
     "LLM system prompt中: '评估是否根据自己的研究思路和逻辑自主撰写（而非套用固定模板）'\n"
     "状态: passed / failed / not_applicable",
     "high", "评估研究内容是否自主撰写而非套模板",
     "MD分章节文本 → LLM"),

    ("LLM-3", "4.1.3 研究基础质量评估", "报告正文（研究基础）", "LLM",
     "LLM system prompt中: '评估是否展示了前期的工作积累'\n"
     "状态: passed / failed / not_applicable",
     "medium", "评估研究基础是否展示前期工作积累",
     "MD分章节文本 → LLM"),
]

REFERENCE_RULES = [
    # L1 级规则 - 基础格式验证
    ("G.L1.01", "文献类型标识检查", "参考文献（每条）", "正则",
     r"re.search(r'\[[A-Z/]+\]', reference)  # 必须有[M],[J]等标识",
     "error", "所有参考文献必须包含文献类型标识",
     "参考文献原文MD"),

    ("G.L1.02", "DOI检查", "参考文献（每条）", "正则",
     r"检测三种DOI格式：URL形式(doi.org) / 字段形式(DOI: 10.xxx) / 直接形式(10.xxxx/xx)",
     "error", "所有参考文献强制要求提供DOI",
     "参考文献原文MD"),

    # L2 级规则 - 完整性检查（需LLM先解析为结构化数据）
    ("G.L2.01", "参考文献解析失败", "参考文献（每条）", "LLM+正则",
     "LLM解析参考文献为结构化字典 → 解析失败则报错",
     "error", "LLM无法将参考文献解析为标准著录项目",
     "LLM结构化JSON"),

    ("G.L2.02", "缺少主要责任者", "参考文献（每条）", "LLM+正则",
     "检查parsed_ref中author字段是否存在",
     "error", "参考文献缺少作者信息（[S]标准类型除外）",
     "LLM结构化JSON"),

    ("G.L2.03", "缺少题名", "参考文献（每条）", "LLM+正则",
     "检查parsed_ref中title字段是否存在",
     "error", "参考文献缺少题名信息",
     "LLM结构化JSON"),

    ("G.L2.04", "缺少出版年", "参考文献（每条）", "LLM+正则",
     "检查parsed_ref中year字段 → 可从原文正则回退提取",
     "error", "参考文献缺少出版年信息",
     "LLM结构化JSON"),

    ("M.L2.01", "专著-缺少出版者", "参考文献（专著）", "LLM+正则",
     "检查parsed_ref中publisher字段",
     "error", "专著类文献缺少出版者",
     "LLM结构化JSON"),

    ("M.L2.02", "专著-缺少出版地", "参考文献（专著）", "LLM+正则",
     "检查parsed_ref中publish_place字段",
     "error", "专著类文献缺少出版地",
     "LLM结构化JSON"),

    ("J.L2.01", "期刊-缺少期刊名", "参考文献（期刊）", "LLM+正则",
     "检查parsed_ref中journal字段",
     "error", "期刊文献缺少期刊名",
     "LLM结构化JSON"),

    ("A.L2.01", "析出文献-缺少出处题名", "参考文献（析出文献）", "LLM+正则",
     "检查parsed_ref中host_title字段",
     "error", "析出文献缺少出处文献题名",
     "LLM结构化JSON"),

    ("A.L2.02", "析出文献-缺少出版地", "参考文献（析出文献）", "LLM+正则",
     "检查parsed_ref中publish_place字段",
     "error", "析出文献缺少出版地",
     "LLM结构化JSON"),

    ("A.L2.03", "析出文献-缺少出版者", "参考文献（析出文献）", "LLM+正则",
     "检查parsed_ref中publisher字段",
     "error", "析出文献缺少出版者",
     "LLM结构化JSON"),

    ("A.L2.04", "析出文献-缺少页码", "参考文献（析出文献）", "LLM+正则",
     "检查parsed_ref中pages字段",
     "error", "析出文献缺少页码信息",
     "LLM结构化JSON"),

    ("N.L2.01", "报纸-缺少报纸名", "参考文献（报纸）", "LLM+正则",
     "检查parsed_ref中newspaper字段",
     "error", "报纸文献缺少报纸名",
     "LLM结构化JSON"),

    ("P.L2.01", "专利-缺少专利号", "参考文献（专利）", "LLM+正则",
     "检查parsed_ref中patent_number字段",
     "error", "专利文献缺少专利号",
     "LLM结构化JSON"),

    ("P.L2.02", "专利-缺少公告日期", "参考文献（专利）", "LLM+正则",
     "检查parsed_ref中announcement_date字段",
     "error", "专利文献缺少公告日期",
     "LLM结构化JSON"),

    ("S.L2.01", "标准-缺少标准编号", "参考文献（标准）", "LLM+正则",
     "检查parsed_ref中standard_number字段",
     "error", "标准文献缺少标准编号",
     "LLM结构化JSON"),

    ("S.L2.02", "标准-缺少出版地", "参考文献（标准）", "LLM+正则",
     "检查parsed_ref中publish_place字段",
     "error", "标准文献缺少出版地",
     "LLM结构化JSON"),

    ("S.L2.03", "标准-缺少出版者", "参考文献（标准）", "LLM+正则",
     "检查parsed_ref中publisher字段",
     "error", "标准文献缺少出版者",
     "LLM结构化JSON"),

    ("D.L2.01", "学位论文-缺少保存单位", "参考文献（学位论文）", "LLM+正则",
     "检查parsed_ref中school字段",
     "error", "学位论文缺少保存单位（学校）",
     "LLM结构化JSON"),

    ("D.L2.02", "学位论文-缺少出版地", "参考文献（学位论文）", "LLM+正则",
     "检查parsed_ref中publish_place字段",
     "error", "学位论文缺少出版地",
     "LLM结构化JSON"),

    ("C.L2.01", "会议论文-缺少出版者", "参考文献（会议论文）", "LLM+正则",
     "检查parsed_ref中publisher字段",
     "error", "会议论文集缺少出版者",
     "LLM结构化JSON"),

    ("C.L2.02", "会议论文-缺少出版地", "参考文献（会议论文）", "LLM+正则",
     "检查parsed_ref中publish_place字段",
     "error", "会议论文集缺少出版地",
     "LLM结构化JSON"),

    ("C.L2.03", "会议论文-缺少出版年", "参考文献（会议论文）", "LLM+正则",
     "检查parsed_ref中year字段",
     "error", "会议论文集缺少出版年",
     "LLM结构化JSON"),

    ("R.L2.01", "报告-缺少出版者", "参考文献（报告）", "LLM+正则",
     "检查parsed_ref中publisher字段",
     "error", "报告文献缺少发布机构",
     "LLM结构化JSON"),

    ("R.L2.02", "报告-缺少出版地", "参考文献（报告）", "LLM+正则",
     "检查parsed_ref中publish_place字段",
     "error", "报告文献缺少出版地",
     "LLM结构化JSON"),

    ("Archive.L2.01", "档案-缺少保存地", "参考文献（档案）", "LLM+正则",
     "检查parsed_ref中archive_place字段",
     "error", "档案文献缺少保存地",
     "LLM结构化JSON"),

    ("Archive.L2.02", "档案-缺少档案馆名称", "参考文献（档案）", "LLM+正则",
     "检查parsed_ref中archive_name字段",
     "error", "档案文献缺少档案馆名称",
     "LLM结构化JSON"),

    ("E.L2.01", "电子资源-缺少URL", "参考文献（电子资源）", "LLM+正则",
     "检查parsed_ref中url字段",
     "error", "电子资源缺少URL",
     "LLM结构化JSON"),

    ("E.L2.02", "电子资源-缺少引用日期", "参考文献（电子资源）", "LLM+正则",
     "检查parsed_ref中access_date字段",
     "error", "电子资源缺少引用日期",
     "LLM结构化JSON"),

    # L3 级规则 - 格式规范性检查
    ("G.L3.01", "作者>3人未用\"等\"", "参考文献（每条）", "LLM+正则",
     "parsed_ref中author_count>3 → 检查原文是否有'等'或'et al.'",
     "info", "作者超过3人时应只列前3人加\"等\"",
     "LLM结构化JSON + 原文MD"),

    ("G.L3.02", "作者<3人误用\"等\"", "参考文献（每条）", "LLM+正则",
     "parsed_ref中author_count≤3 → 检查原文不应有'等'或'et al.'",
     "info", "作者不足3人不应使用\"等\"",
     "LLM结构化JSON + 原文MD"),

    ("G.L3.03", "汉语拼音姓名格式不规范", "参考文献（每条）", "LLM+正则",
     r"re.search拼音姓名模式 → 姓氏需全大写(如 ZHANG S)",
     "info", "汉语拼音姓氏应全大写",
     "LLM结构化JSON + 原文MD"),

    ("G.L3.04", "其他责任者著录不规范", "参考文献（每条）", "LLM+正则",
     "检查'译','编','注'等责任者标识的位置和格式",
     "info", "译/编/注等其他责任者著录是否规范",
     "LLM结构化JSON + 原文MD"),

    ("G.L3.05", "副标题分隔符格式不规范", "参考文献（每条）", "LLM+正则",
     r"题名中副标题分隔符应为': '（冒号+空格）",
     "info", "副标题分隔符应为冒号加空格",
     "LLM结构化JSON + 原文MD"),

    ("G.L3.06", "引文页码分隔符不规范", "参考文献（每条）", "LLM+正则",
     "引文页码前分隔符应为':'而非','",
     "info", "引文页码前应使用冒号分隔",
     "LLM结构化JSON + 原文MD"),

    ("G.L3.07", "标点符号使用不规范", "参考文献（每条）", "正则",
     r"validate_punctuation: 检查全角/半角标点、作者与题名间句号、多作者间逗号等",
     "info", "标点符号使用需符合GB/T 7714规范",
     "参考文献原文MD"),

    ("M.L3.01", "专著-版本项著录不规范", "参考文献（专著）", "LLM+正则",
     "第1版不需著录；其他版次格式应为'第X版'或'Xth ed.'",
     "info", "第1版不应标注；其他版次格式需规范",
     "LLM结构化JSON + 原文MD"),

    ("M.L3.02", "专著-出版信息不详未用方括号", "参考文献（专著）", "LLM+正则",
     "出版地不详用[S.l.]，出版者不详用[s.n.]",
     "info", "出版地/出版者不详时需方括号标识",
     "LLM结构化JSON + 原文MD"),

    ("A.L3.01", "析出文献分隔符错误", "参考文献（析出文献）", "LLM+正则",
     r"专著用'//'分隔，期刊用'.'分隔",
     "info", "析出文献与出处间分隔符需正确",
     "LLM结构化JSON + 原文MD"),

    ("A.L3.02", "析出文献页码格式不规范", "参考文献（析出文献）", "LLM+正则",
     "页码格式应为 起页-终页",
     "info", "析出文献页码格式需规范",
     "LLM结构化JSON + 原文MD"),

    ("J.L3.01", "期刊年卷期页码格式不规范", "参考文献（期刊）", "LLM+正则",
     r"格式: 年, 卷(期): 页码  如 2020, 45(3): 100-110",
     "info", "年/卷/期/页码格式需规范",
     "LLM结构化JSON + 原文MD"),

    ("J.L3.02", "期刊题名缩写不符合ISO 4", "参考文献（期刊）", "LLM+正则",
     "启发式检查期刊名是否使用了非标准缩写",
     "info", "期刊名缩写应符合ISO 4规范",
     "LLM结构化JSON + 原文MD"),

    ("N.L3.01", "报纸出版日期版次格式不规范", "参考文献（报纸）", "LLM+正则",
     "格式应为 YYYY-MM-DD(版次)",
     "info", "报纸日期和版次格式需规范",
     "LLM结构化JSON + 原文MD"),

    ("N.L3.02", "报纸版次著录格式不正确", "参考文献（报纸）", "LLM+正则",
     "版次应在圆括号中",
     "info", "报纸版次著录格式需规范",
     "LLM结构化JSON + 原文MD"),

    ("P.L3.01", "专利公告日期格式不正确", "参考文献（专利）", "LLM+正则",
     "格式应为 YYYY-MM-DD",
     "info", "专利公告日期应为YYYY-MM-DD格式",
     "LLM结构化JSON + 原文MD"),

    ("P.L3.02", "专利号与题名分隔符不正确", "参考文献（专利）", "LLM+正则",
     r"分隔符应为':'",
     "info", "专利号与题名间应用冒号分隔",
     "LLM结构化JSON + 原文MD"),

    ("S.L3.01", "标准编号格式不规范", "参考文献（标准）", "LLM+正则",
     "如 GB/T 7714-2015 格式",
     "info", "标准编号格式或位置需规范",
     "LLM结构化JSON + 原文MD"),

    ("D.L3.01", "学位论文保存地格式不规范", "参考文献（学位论文）", "LLM+正则",
     "格式应为 地点：学校",
     "info", "保存地与保存单位格式需规范",
     "LLM结构化JSON + 原文MD"),

    ("C.L3.01", "会议论文格式不规范", "参考文献（会议论文）", "LLM+正则",
     r"缺少'//'分隔符或页码格式问题",
     "info", "会议论文需'//'分隔符和正确页码格式",
     "LLM结构化JSON + 原文MD"),

    ("C.L3.02", "会议论文集编者信息缺失", "参考文献（会议论文）", "LLM+正则",
     "编者信息位置应在会议名后",
     "info", "会议论文集编者信息需正确著录",
     "LLM结构化JSON + 原文MD"),

    ("C.L3.03", "会议日期格式不规范", "参考文献（会议论文）", "LLM+正则",
     "会议日期格式需规范",
     "info", "会议日期格式需规范",
     "LLM结构化JSON + 原文MD"),

    ("C.L3.04", "建议包含会议名称", "参考文献（会议论文）", "LLM+正则",
     "会议论文建议注明会议名称",
     "info", "建议包含会议名称信息",
     "LLM结构化JSON + 原文MD"),

    ("C.L3.05", "会议地点格式不规范", "参考文献（会议论文）", "LLM+正则",
     "会议地点应在会议名后",
     "info", "会议地点格式需规范",
     "LLM结构化JSON + 原文MD"),

    ("C.L3.06", "会议论文建议包含页码", "参考文献（会议论文）", "LLM+正则",
     "建议添加页码信息",
     "info", "建议包含页码信息",
     "LLM结构化JSON + 原文MD"),

    ("R.L3.01", "报告编号格式不规范", "参考文献（报告）", "LLM+正则",
     "报告编号位置和格式需规范",
     "info", "报告编号格式或位置不正确",
     "LLM结构化JSON + 原文MD"),

    ("E.L3.01", "电子资源引用日期格式不正确", "参考文献（电子资源）", "LLM+正则",
     "格式应为 [YYYY-MM-DD]",
     "info", "引用日期需方括号且格式为YYYY-MM-DD",
     "LLM结构化JSON + 原文MD"),

    ("E.L3.02", "电子资源载体类型标识不规范", "参考文献（电子资源）", "LLM+正则",
     "应为[OL],[CD],[MT]等",
     "info", "载体类型标识需规范",
     "LLM结构化JSON + 原文MD"),

    ("Archive.L3.01", "档案标识符缺失", "参考文献（档案）", "LLM+正则",
     "[A]标识缺失或位置不对",
     "info", "档案文献需有[A]标识",
     "LLM结构化JSON + 原文MD"),

    ("Archive.L3.02", "档案号格式不规范", "参考文献（档案）", "LLM+正则",
     "档案号格式需规范",
     "info", "档案号格式不规范",
     "LLM结构化JSON + 原文MD"),
]


# ==================== Excel 生成逻辑 ====================

HEADERS = ["序号", "规则名称", "检查位置", "实现方式", "正则/LLM提示词", "严重度", "简要说明", "数据来源"]

# 列宽设置
COL_WIDTHS = [8, 32, 24, 16, 60, 12, 40, 30]

# 颜色主题
COLORS = {
    'header_fill': PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid'),
    'header_font': Font(color='FFFFFF', bold=True, size=11, name='微软雅黑'),
    'category_fill': PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),
    'category_font': Font(bold=True, size=11, name='微软雅黑', color='1F4E79'),
    'data_font': Font(size=10, name='微软雅黑'),
    'severity_high': Font(size=10, name='微软雅黑', color='CC0000', bold=True),
    'severity_medium': Font(size=10, name='微软雅黑', color='E68A00'),
    'severity_low': Font(size=10, name='微软雅黑', color='336699'),
    'severity_error': Font(size=10, name='微软雅黑', color='CC0000', bold=True),
    'severity_info': Font(size=10, name='微软雅黑', color='666666'),
}

THIN_BORDER = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)


def _get_severity_font(severity: str) -> Font:
    """根据严重度返回字体颜色"""
    s = severity.lower()
    if 'error' in s or 'high' in s:
        return COLORS['severity_high']
    elif 'medium' in s:
        return COLORS['severity_medium']
    elif 'low' in s:
        return COLORS['severity_low']
    elif 'info' in s:
        return COLORS['severity_info']
    return COLORS['data_font']


def _write_header(ws, row: int):
    """写入表头"""
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = COLORS['header_fill']
        cell.font = COLORS['header_font']
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    return row + 1


def _write_category(ws, row: int, category_name: str, col_count: int):
    """写入分类标题行"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=category_name)
    cell.fill = COLORS['category_fill']
    cell.font = COLORS['category_font']
    cell.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(1, col_count + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER
    return row + 1


def _write_rules(ws, row: int, rules: list):
    """写入规则数据行"""
    for rule in rules:
        for col_idx, value in enumerate(rule, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = COLORS['data_font']
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = THIN_BORDER
            # 严重度列特殊着色
            if col_idx == 6:
                cell.font = _get_severity_font(str(value))
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # 序号列居中
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            # 实现方式列居中
            if col_idx == 4:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # 数据来源列居中
            if col_idx == 8:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row += 1
    return row


def _build_sheet(ws, doc_type: str):
    """构建一张Sheet的完整内容

    doc_type:
        'PDF'  — NSFC国家自然科学基金申请书，包含全部 87 条规则
        'Word' — 大学生毕业论文，仅包含 60 条参考文献 GB/T 7714 验证规则
    """
    ws.sheet_properties.tabColor = '1F4E79' if doc_type == 'PDF' else '2E75B6'

    # 设置列宽
    for idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    col_count = len(HEADERS)

    # ---------- 标题行 ----------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    if doc_type == 'PDF':
        total_pdf = len(PRELIMINARY_RULES) + len(LLM_QUALITY_RULES) + len(REFERENCE_RULES)
        title_text = f"NSFC申请书检查规则汇总（PDF版）— 共 {total_pdf} 条规则"
    else:
        title_text = f"大学生毕业论文检查规则汇总（Word版）— 共 {len(REFERENCE_RULES)} 条规则"
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=14, name='微软雅黑', color='1F4E79')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    for c in range(1, col_count + 1):
        ws.cell(row=1, column=c).border = THIN_BORDER

    # ---------- 副标题说明 ----------
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    if doc_type == 'PDF':
        sub_text = ("检查对象：NSFC国家自然科学基金申请书  |  "
                    "文档来源：PDF文件 → MinerU转Markdown → VLM表格提取 + 正则初筛 + LLM质量评估 + 参考文献验证")
    else:
        sub_text = ("检查对象：大学生毕业论文  |  "
                    "文档来源：Word文件 → python-docx提取参考文献 → GB/T 7714参考文献验证")
    sub_cell = ws.cell(row=2, column=1, value=sub_text)
    sub_cell.font = Font(size=9, name='微软雅黑', color='666666')
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 4

    if doc_type == 'PDF':
        # ====== PDF版：三部分全部输出 ======

        # 第一部分：初筛规则
        row = _write_category(ws, row, f"一、初筛规则（正则/混合）— 共 {len(PRELIMINARY_RULES)} 条", col_count)
        row = _write_header(ws, row)
        row = _write_rules(ws, row, PRELIMINARY_RULES)
        row += 1

        # 第二部分：LLM 正文质量评估
        row = _write_category(ws, row, f"二、LLM正文质量评估规则 — 共 {len(LLM_QUALITY_RULES)} 条", col_count)
        row = _write_header(ws, row)
        row = _write_rules(ws, row, LLM_QUALITY_RULES)
        row += 1

        # 第三部分：参考文献 GB/T 7714 验证
        row = _write_category(ws, row, f"三、参考文献GB/T 7714验证规则 — 共 {len(REFERENCE_RULES)} 条", col_count)
        row = _write_header(ws, row)
        row = _write_rules(ws, row, REFERENCE_RULES)
        row += 1

        # 汇总行
        total = len(PRELIMINARY_RULES) + len(LLM_QUALITY_RULES) + len(REFERENCE_RULES)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
        summary_cell = ws.cell(row=row, column=1,
                               value=f"合计：{total} 条规则  "
                                     f"（初筛 {len(PRELIMINARY_RULES)} + LLM质量 {len(LLM_QUALITY_RULES)} "
                                     f"+ 参考文献验证 {len(REFERENCE_RULES)}）")
    else:
        # ====== Word版：仅参考文献验证 ======

        row = _write_category(ws, row, f"参考文献GB/T 7714验证规则 — 共 {len(REFERENCE_RULES)} 条", col_count)
        row = _write_header(ws, row)
        row = _write_rules(ws, row, REFERENCE_RULES)
        row += 1

        # 汇总行
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
        summary_cell = ws.cell(row=row, column=1,
                               value=f"合计：{len(REFERENCE_RULES)} 条参考文献验证规则")

    summary_cell.font = Font(bold=True, size=11, name='微软雅黑', color='1F4E79')
    summary_cell.alignment = Alignment(horizontal='center', vertical='center')
    for c in range(1, col_count + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER

    # 冻结窗格（冻结标题和表头）
    ws.freeze_panes = 'A4'


def generate_rules_excel(output_path: str = None):
    """生成检查规则Excel文件"""
    if output_path is None:
        output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '检查规则汇总.xlsx')

    wb = Workbook()

    # Sheet 1: PDF版检查规则（NSFC申请书）
    ws_pdf = wb.active
    ws_pdf.title = "PDF版-NSFC申请书"
    _build_sheet(ws_pdf, "PDF")

    # Sheet 2: Word版检查规则（大学生毕业论文）
    ws_word = wb.create_sheet("Word版-毕业论文")
    _build_sheet(ws_word, "Word")

    wb.save(output_path)
    total_pdf = len(PRELIMINARY_RULES) + len(LLM_QUALITY_RULES) + len(REFERENCE_RULES)
    print(f"✅ 规则Excel已生成: {output_path}")
    print(f"   === PDF版（NSFC申请书）===")
    print(f"   初筛规则: {len(PRELIMINARY_RULES)} 条")
    print(f"   LLM质量规则: {len(LLM_QUALITY_RULES)} 条")
    print(f"   参考文献验证规则: {len(REFERENCE_RULES)} 条")
    print(f"   PDF版合计: {total_pdf} 条")
    print(f"   === Word版（毕业论文）===")
    print(f"   参考文献验证规则: {len(REFERENCE_RULES)} 条")
    print(f"   Word版合计: {len(REFERENCE_RULES)} 条")

    return output_path


if __name__ == '__main__':
    generate_rules_excel()
